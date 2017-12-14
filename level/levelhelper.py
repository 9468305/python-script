#!/usr/local/bin/python3
# -*- coding: utf-8 -*-
'''leveldb helper'''

import leveldb
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl import cell


def exist(db_src, key):
    '''Safely check whether key exist or not'''
    _key_obj = bytes(key.encode('utf-8')) if isinstance(key, str) else key
    try:
        db_src.Get(_key_obj)
        return True
    except KeyError:
        return False


def count(db_src, k_filter, v_filter):
    '''Count database items, support key filter and/or value filter, return total and valid.'''
    total, valid = 0, 0
    for _k, _v in db_src.RangeIter():
        total += 1
        if k_filter:
            if _k.find(k_filter) == -1:
                continue
        if v_filter:
            if _v.find(v_filter) == -1:
                continue
        valid += 1
    return total, valid


def copy(db_src, db_dst, k_filter):
    '''copy db_src to db_dst, support key filter, return total and valid.'''
    total, valid = 0, 0
    for _k, _v in db_src.RangeIter():
        total += 1
        if k_filter:
            if _k.find(k_filter) != -1:
                valid += 1
                db_dst.Put(_k, _v, sync=True)
        else:
            valid += 1
            db_dst.Put(_k, _v, sync=True)
    return total, valid


def delete(db_src, db_dst):
    '''Delete db_src items in db_dst.'''
    for _k, _v in db_src.RangeIter():
        db_dst.Delete(_k)


def diff(db_src, db_dst, db_diff):
    '''Find differents between db_src and db_dst, save to db_diff, return diff count.'''
    diff_count = 0
    for _k, _v in db_src.RangeIter():
        if not exist(db_dst, _k):
            diff_count += 1
            db_diff.Put(_k, _v)
    return diff_count


def clean_copy(db_src, db_dst):
    '''copy db_src to db_dst, clean empty value, return total count.'''
    total = 0
    for _k, _v in db_src.RangeIter():
        if _v:
            db_dst.Put(_k, _v)
            total += 1
    return total


def dump(db_src):
    '''Dump database key and value items.'''
    _db = leveldb.LevelDB(db_src, create_if_missing=False) if isinstance(db_src, str) else db_src
    for _k, _v in _db.RangeIter():
        print(_k.decode(), _v.decode())


def db_to_text(from_db, to_text):
    '''Transfer leveldb to text file.'''
    _db = leveldb.LevelDB(from_db, create_if_missing=False) if isinstance(from_db, str) else from_db
    with open(to_text, 'w', encoding='utf-8') as _f:
        for _k, _v in _db.RangeIter():
            _f.write(_k.decode() + ',' + _v.decode() + '\n')


def text_to_db(from_text, to_db, split_char):
    '''Transfer text file to leveldb, return total and invalid count.'''
    total, invalid = 0, 0
    _split = split_char if split_char else ','
    _db = leveldb.LevelDB(to_db, create_if_missing=True) if isinstance(to_db, str) else to_db
    with open(from_text, 'r', encoding='utf-8') as _f:
        lines = _f.readlines()
        total = len(lines)
        for line in lines:
            if not line:
                invalid += 1
                continue
            # line = line.strip()
            if _split in line:
                _sub = line.split(_split, 1)
                _db.Put(_sub[0].encode('utf-8'), _sub[1].encode('utf-8'))
            else:
                _db.Put(line, '')
        return total, invalid


def db_to_excel(from_db, to_excel):
    '''Transfer leveldb to Excel file, return total count.'''
    _db = leveldb.LevelDB(from_db, create_if_missing=False) if isinstance(from_db, str) else from_db
    _wb = Workbook()
    _ws = _wb.active
    total = 0
    for _k, _v in _db.RangeIter():
        _ws.append([_k.decode(), _v.decode()])
        total += 1
    _wb.save(to_excel)
    return total


def excel_to_db(from_excel, to_db):
    '''Transfer Excel file to leveldb, return total count.'''
    _wb = load_workbook(from_excel, read_only=True)
    _ws = _wb.active
    _db = leveldb.LevelDB(to_db, create_if_missing=True) if isinstance(to_db, str) else to_db
    total = 0
    for _row in _ws.iter_rows(min_row=2, min_col=1, max_col=1):
        if _row and _row[0] and _row[1]:
            _key, _value = '', ''
            if _row[0].data_type == cell.Cell.TYPE_STRING:
                _key = _row[0].value.encode('utf-8')
                _key = ''.join(_key.split())
            if _row[1].data_type == cell.Cell.TYPE_STRING:
                _value = _row[0].value.encode('utf-8')
                _value = ''.join(_value.split())
            _db.Put(_key, _value)
            total += 1

    _wb.close()
    return total
