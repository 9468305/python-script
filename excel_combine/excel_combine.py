#!/usr/local/bin/python3
# -*- coding: utf-8 -*-
'''合并指定文件夹下所有Excel文件到同一个文件'''
import os
import collections
import operator
from openpyxl import load_workbook
from openpyxl import Workbook

def search_excel(from_dir, to_file):
    '''遍历from_dir文件夹，查找Excel文件，返回结果列表'''
    _results = []
    for _root, _dirs, _files in os.walk(from_dir):
        for _file in _files:
            if _file.endswith('.xlsx'):
                _results.append(os.path.join(_root, _file))

    try:
        print('Remove combine.xlsx.')
        _results.remove(to_file)
    except ValueError:
        print('combine.xlsx not exist.')
    return _results


def load_excel(excel_file):
    '''读取Excel文件内容，返回Excel的标题数组和数据有序字典'''
    _wb = load_workbook(excel_file, read_only=True)
    _ws = _wb.active
    _title = []
    _items = collections.OrderedDict()
    for _r in _ws.rows:
        if not _title:
            for _i in _r:
                _title.append(_i.value)
        else:
            _item = []
            for _i in _r:
                _item.append(_i.value)
            _items[_item[0]] = _item

    _wb.close()
    return _title, _items


def save_excel(excel_file, excel_title, excel_items):
    '''保存Excel文件'''
    _wb = Workbook()
    _ws = _wb.active
    _ws.append(excel_title)
    for _k, _v in excel_items.items():
        _ws.append(_v)
    _wb.save(excel_file)


def combine(from_dir, to_file):
    '''合并指定文件夹下所有Excel文件到同一个文件'''
    _excel_files = search_excel(from_dir, to_file)
    if not _excel_files:
        return
    _excel_title = []
    _excel_content = collections.OrderedDict()
    for _file in _excel_files:
        print('Parsing ' + _file)
        _title, _items = load_excel(_file)
        if not _title or not _items:
            print('Skip since it is empty.')
            continue

        if not _excel_title:
            _excel_title = _title
        elif not operator.eq(_title, _excel_title):
            print('Warning: Excel title format are different!')

        for _k, _v in _items.items():
            _excel_content[_k] = _v
        print('Parsing done.')

    if not _excel_title or not _excel_content:
        print('All files is empty.')
        return
    save_excel(to_file, _excel_title, _excel_content)


if __name__ == "__main__":
    print('begin')
    FROM_DIR = os.getcwd()
    TO_FILE = os.path.join(FROM_DIR, 'combine.xlsx')
    combine(FROM_DIR, TO_FILE)
    print('end')
