#!/usr/local/bin/python3
# -*- coding: utf-8 -*-
'''
geetest offline 6.0.0 spider for gsxt 广东
'''

import os
import time
import random
import logging
import traceback
import json
import requests
import execjs
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl import Workbook
import constants
import util

# Set default logging handler to avoid "No handler found" warnings.
try:  # Python 2.7+
    from logging import NullHandler
except ImportError:
    class NullHandler(logging.Handler):
        '''NullHandler'''
        def emit(self, record):
            pass

logging.getLogger(__name__).addHandler(NullHandler())
logging.basicConfig(level=logging.INFO)

HOST = 'http://gd.gsxt.gov.cn'
INDEX = HOST

JSRUNTIME = execjs.get(execjs.runtime_names.Node)

USERRESPONSE_JSCONTEXT = JSRUNTIME.compile(util.USERRESPONSE_JS)

TIMEOUT = 15

def load_lists_excel(excel_file):
    '''加载lists excel文件，仅有一列公司名称数据，作为查询keyword。'''
    _items = []
    if not os.path.isfile(excel_file):
        logging.info("Excel File Not Exist")
        return _items

    _wb = load_workbook(excel_file, read_only=True)
    _ws = _wb.active
    _title = None

    for _r in _ws.rows:
        if not _title:
            _title = _r[0].value
        else:
            _items.append(_r[0].value)

    _wb.close()
    logging.debug(_title)
    logging.debug(_items)
    logging.debug(len(_items))
    return _items


def save_result_excel(excel_file, results):
    '''save result excel'''
    _wb = Workbook()
    _ws = _wb.active
    _ws.append(['企业名称',
                '注册号/统一社会信用代码',
                '类型',
                '法定代表人',
                '注册资本',
                '成立日期',
                '营业期限自',
                '营业期限至',
                '登记机关',
                '核准日期',
                '登记状态',
                '住所',
                '经营范围'
               ])
    for _r in results:
        _ws.append([_r['企业名称'],
                    _r['注册号/统一社会信用代码'],
                    _r['类型'],
                    _r['法定代表人'],
                    _r['注册资本'],
                    _r['成立日期'],
                    _r['营业期限自'],
                    _r['营业期限至'],
                    _r['登记机关'],
                    _r['核准日期'],
                    _r['登记状态'],
                    _r['住所'],
                    _r['经营范围']
                   ])
    _wb.save(excel_file)


def load_json(json_file):
    '''load json file'''
    if not os.path.isfile(json_file):
        logging.info("Json File Not Exist")
        return None

    with open(json_file, 'r', encoding='utf8') as _f:
        json_data = json.load(_f)
        logging.info(len(json_data))
        return json_data


def save_json(json_file, json_data):
    '''save json file'''
    with open(json_file, 'w', encoding='utf8') as _f:
        json.dump(json_data, _f, indent=2, sort_keys=True, ensure_ascii=False)
    logging.info(len(json_data))


def excel_to_json(excel_file, json_file):
    '''convert excel file to json file'''
    lists = load_lists_excel(excel_file)
    logging.info(len(lists))
    save_json(json_file, lists)


def json_to_excel(json_file, excel_file):
    '''convert json file to excel file'''
    print('todo')


def calc_userresponse(distance, challenge):
    '''根据滑动距离distance和challenge，计算userresponse值'''
    return USERRESPONSE_JSCONTEXT.call('userresponse', distance, challenge)


def calc_validate(challenge):
    '''计算validate值'''
    _r = random.randint(0, len(util.OFFLINE_SAMPLE)-1)
    distance, rand0, rand1 = util.OFFLINE_SAMPLE[_r]
    distance_r = calc_userresponse(distance, challenge)
    rand0_r = calc_userresponse(rand0, challenge)
    rand1_r = calc_userresponse(rand1, challenge)
    validate = distance_r + '_' + rand0_r + '_' + rand1_r
    logging.debug(validate)
    return validate


def parse_name_url(html_doc):
    '''使用BeautifulSoup解析HTML页面，查找详情链接'''
    _soup = BeautifulSoup(html_doc, 'html.parser')
    _findall = _soup.find_all('div',
                              class_="clickStyle",
                              style='margin-left: 160px;padding-left: 10px;')
    name_url_array = []
    if _findall:
        for _a in _findall:
            _company = _a.find('a')
            _name = ''.join(_company.get_text().split())
            _url = _company['href']
            if _url.startswith('../'):
                _url = INDEX + '/aiccips/CheckEntContext/' + _url
            name_url_array.append((_name, _url))
        logging.info(name_url_array)
    else:
        logging.error('Company Link Not Found')
    return name_url_array


def get_mainpage(session):
    '''Get mainpage'''
    logging.debug('GET ' + INDEX)
    _headers = {'Accept': constants.ACCEPT_HTML,
                'Accept-Language': constants.ACCEPT_LANGUAGE,
                'User-Agent': constants.USER_AGENT}
    _response = session.get(INDEX, headers=_headers, timeout=TIMEOUT)
    logging.debug('response code:' + str(_response.status_code))
    return _response.status_code == 200


def get_captcha(session):
    '''
    GET /aiccips//verify/start.html
    Response
    {
	    "success": 0,
	    "gt": "c02ee51ee0afe88899efe6dc729627fc",
	    "challenge": "ed3d2c21991e3bef5e069713af9fa6caed"
    }
    '''
    _url = INDEX + '/aiccips//verify/start.html'
    logging.debug('GET ' + _url)
    _headers = {'Accept': constants.ACCEPT_JSON,
                'Accept-Language': constants.ACCEPT_LANGUAGE,
                'User-Agent': constants.USER_AGENT,
                'Referer': INDEX,
                'X-Requested-With': 'XMLHttpRequest'}
    _params = {'t': str(int(time.time() * 1000))}
    _response = session.get(_url, headers=_headers, params=_params, timeout=TIMEOUT)
    logging.debug('response code: ' + str(_response.status_code))
    logging.debug('response text: ' + _response.text)
    if _response.status_code != 200:
        return False
    return _response.json()


def post_validate(session, challenge, validate, keyword):
    '''
    POST /aiccips/verify/sec.html
    Response
    {
	    "status": "success",
	    "textfield": "waY5F5lZyxvKw9bMM4nBs7HUgWS1SRpagFutRKqs/+DkRqCIS9N4PUCqM9fmrbg1",
	    "version": "3.3.0"
    }
    '''
    _url = INDEX + '/aiccips/verify/sec.html'
    logging.debug('POST ' + _url)
    _headers = {'Accept': constants.ACCEPT_JSON,
                'Accept-Language': constants.ACCEPT_LANGUAGE,
                'User-Agent': constants.USER_AGENT,
                'Referer': INDEX,
                'X-Requested-With': 'XMLHttpRequest',
                'Origin': HOST}
    _params = [('textfield', keyword),
               ('geetest_challenge', challenge),
               ('geetest_validate', validate),
               ('geetest_seccode', validate + '|jordan')]
    _response = session.post(_url, headers=_headers, data=_params, timeout=TIMEOUT)
    logging.debug('response code: ' + str(_response.status_code))
    logging.debug('response text: ' + _response.text)
    if _response.status_code != 200:
        return False
    _json_obj = _response.json()
    logging.debug(_json_obj)
    return _json_obj['textfield'] if _json_obj['status'] == 'success' else None


def post_search(session, textfield):
    '''POST /aiccips/CheckEntContext/showCheck.html'''
    _url = INDEX + '/aiccips/CheckEntContext/showCheck.html'
    logging.debug('POST ' + _url)
    _headers = {'Accept': constants.ACCEPT_HTML,
                'Accept-Language': constants.ACCEPT_LANGUAGE,
                'User-Agent': constants.USER_AGENT,
                'Referer': INDEX,
                'X-Requested-With': 'XMLHttpRequest',
                'Origin': HOST}
    _params = [('textfield', textfield),
               ('type', 'nomal')]
    _response = session.post(_url, headers=_headers, data=_params, timeout=TIMEOUT)
    logging.debug('response code: ' + str(_response.status_code))
    logging.debug('response text: ' + _response.text)
    if _response.status_code != 200:
        return None
    return parse_name_url(_response.text)


def get_validate(session, keyword):
    '''循环进行validate验证'''
    for _ in range(10):
        captcha = get_captcha(session)
        if not captcha:
            return None

        validate = calc_validate(captcha['challenge'])
        textfield = post_validate(session, captcha['challenge'], validate, keyword)
        if textfield:
            return textfield
    return None


def parse_detail(html_doc):
    '''parse company detail'''
    _soup = BeautifulSoup(html_doc, 'html.parser')
    _yyzz = _soup.find('div', class_='item_box', id='yyzz')
    if not _yyzz:
        logging.error('Detail yyzz Not Found')
        return None

    _li_all = _yyzz.find_all('li')
    if not _li_all:
        logging.error("Detail li Not Found")
        return None

    _info = {}
    for _li in _li_all:
        _text = ''.join(_li.get_text().split())
        _k, _v = _text.split(sep='：', maxsplit=1)
        _info[_k] = _v
    logging.info(_info)
    if not _info['企业名称']:
        _info = None # for safe
    return _info


def parse_detail_2(html_doc):
    '''parse company detail 2'''
    _soup = BeautifulSoup(html_doc, 'html.parser')
    _table = _soup.find('table', cellspacing='6')
    if not _table:
        logging.error('Detail table Not Found')
        return None

    _tr_all = _table.find_all('td')
    if not _tr_all:
        logging.error("Detail td Not Found")
        return None

    _info = {}
    for _td in _tr_all:
        _text = ''.join(_td.get_text().split())
        if _text == '营业执照信息':
            continue
        _k, _v = _text.split(sep='：', maxsplit=1)
        _temp = {}
        _temp[_k] = _v
        for _k2, _v2 in _temp.items():
            if _k2 == '.企业名称' or _k2 == '.名称':
                _info['企业名称'] = _v2
            elif _k2 == '.统一社会信用代码/注册号' or _k2 == '.注册号':
                _info['注册号/统一社会信用代码'] = _v2
            elif _k2 == '.类型':
                _info['类型'] = _v2
            elif _k2 == '.负责人' or _k2 == '.经营者':
                _info['法定代表人'] = _v2
            elif _k2 == '.成立日期' or _k2 == '.注册日期':
                _info['成立日期'] = _v2
            elif _k2 == '.营业期限自':
                _info['营业期限自'] = _v2
            elif _k2 == '.营业期限至':
                _info['营业期限至'] = _v2
            elif _k2 == '.登记机关':
                _info['登记机关'] = _v2
            elif _k2 == '.核准日期':
                _info['核准日期'] = _v2
            elif _k2 == '.登记状态':
                _info['登记状态'] = _v2
            elif _k2 == '.营业场所' or _k2 == '.经营场所':
                _info['住所'] = _v2
            elif _k2 == '.经营范围':
                _info['经营范围'] = _v2
        _info['注册资本'] = '0'
    logging.info(_info)
    if not _info['企业名称']:
        _info = None # for safe
    return _info


def query_keyword(session, keyword):
    '''query keyword'''
    #if not get_mainpage(session):
    #    return None
    logging.info(keyword)
    textfield = get_validate(session, keyword)
    if textfield:
        return post_search(session, textfield)
    return None


def safe_query_keyword(keyword):
    '''Safe query keyword, handle network timeout and retry'''
    for _ in range(5):
        try:
            with requests.Session() as session:
                return query_keyword(session, keyword)
        except requests.RequestException as _e:
            traceback.print_exc()
            time.sleep(5)
    return None


def query_detail(session, url):
    '''query company detail url'''
    logging.debug('GET ' + url)
    _headers = {'Accept': constants.ACCEPT_HTML,
                'Accept-Language': constants.ACCEPT_LANGUAGE,
                'User-Agent': constants.USER_AGENT}
    _response = session.get(url, headers=_headers, timeout=TIMEOUT)
    logging.debug('response code:' + str(_response.status_code))
    if _response.status_code == 200:
        if url.find('www.szcredit.org.cn') is not -1:
            return parse_detail(_response.text)
        elif url.find('GSpublicityList.html') is not -1:
            return parse_detail_2(_response.text)
        else:
            logging.error('URL Type Not Support')
    return None


def safe_query_detail(url):
    '''Safe query url, handle network timeout and retry'''
    for _ in range(5):
        try:
            with requests.Session() as session:
                return query_detail(session, url)
        except requests.RequestException as _e:
            traceback.print_exc()
            time.sleep(5)
    return None


def main():
    '''main entry'''
    lists = load_json('list.json')
    if not lists:
        lists = []
    results = load_json('result.json')
    if not results:
        results = {}
    notfound = load_json('notfound.json')
    if not notfound:
        notfound = []

    for keyword in lists:
        if keyword in results:
            continue
        if keyword in notfound:
            continue
        name_url_array = safe_query_keyword(keyword)
        if not name_url_array:
            notfound.append(keyword)
            continue
        for name, url in name_url_array:
            if name in results:
                continue
            detail_dict = safe_query_detail(url)
            if detail_dict:
                results.update({name : detail_dict})
    save_json('result.json', results)
    save_json('notfound.json', notfound)
    logging.info('done')


if __name__ == "__main__":
    #excel_to_json('list.xlsx', 'list.json')
    #json_to_excel('result.json', 'result.xlsx')
    main()
